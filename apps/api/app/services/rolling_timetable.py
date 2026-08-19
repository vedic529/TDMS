"""Rolling timetable workbook parser and normaliser.

The operational rolling timetable is maintained as one worksheet per
qualification. Each row is a calendar week; each column from the fourth onward
is one rolling intake, and a cell says what that intake is doing that week.

    Week No. | Start Date | End Date | <intake> | <intake> | ...
           1 | 19-Jan-26  | 25-Jan-26| BSBCRT511|          |
           3 | 02-Feb-26  | 08-Feb-26| BSBCRT511| Intake   |
           4 | 09-Feb-26  | 15-Feb-26| BSBLDR523| BSBLDR523|

**This is a rolling timetable, not a fixed sequence.** Intakes join the cycle at
different points and wrap around: the cohort marked at week 3 above begins with
BSBLDR523, not with the first unit of the cycle, and meets BSBCRT511 a year
later when the cycle comes round. Anything that assumes every cohort starts at
unit one will produce a wrong timetable.

Two decisions worth stating, because both are easy to get subtly wrong:

* **Nothing here is keyed to a spreadsheet coordinate.** Columns are found by
  scanning, not by naming `E4`. Adding an intake column or a calendar year to
  the workbook does not require touching this module.
* **No unit code is hard-coded.** The parser reports the codes it read; whether
  they are units of the qualification is decided against the database by the
  caller, so the workbook cannot introduce a unit that TDMS has not approved.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field
from itertools import groupby
from typing import Iterable, Sequence

#: Cell values that schedule something other than a unit of competency.
#:
#: A break is not a unit: it has no trainer, no qualification membership and is
#: not one of the qualification's units. Same for an assessment week. Both are
#: course-calendar events, and creating a `Unit` row with the code `BREAK` would
#: corrupt the national unit register.
BREAK = "BREAK"
ASSESSMENT = "ASSESSMENT WEEK"
INTAKE_MARKER = "INTAKE"

NON_UNIT_VALUES = {BREAK, ASSESSMENT, INTAKE_MARKER}

PERIOD_UNIT = "UNIT"
PERIOD_BREAK = "BREAK"
PERIOD_ASSESSMENT = "ASSESSMENT"

_HEADER_WEEK = "week"
_HEADER_START = "start"
_HEADER_END = "end"


class RollingTimetableError(Exception):
    """The workbook does not have the structure this parser requires."""


@dataclass(frozen=True)
class CalendarWeek:
    """One row: a numbered week and the dates it covers."""

    week_no: int
    start_date: dt.date
    end_date: dt.date


@dataclass
class Period:
    """A contiguous run of identical weeks within one intake column.

    `active_weeks` is the number of calendar weeks in this run. It is not the
    same as the span of a unit's delivery when a break interrupts it — see
    :class:`UnitDelivery`.
    """

    period_type: str
    value: str | None
    first_week_no: int
    last_week_no: int
    start_date: dt.date
    end_date: dt.date
    active_weeks: int
    #: Where this came from, for traceability back to the workbook.
    source_column: str = ""


@dataclass
class UnitDelivery:
    """One unit's delivery for one intake, breaks included rather than removed.

    A unit interrupted by a break is **one delivery**, not two. The workbook
    shows BSBPEF502 for two weeks, then four weeks of break, then BSBPEF502
    again; that is a single delivery of three teaching weeks spanning seven
    calendar weeks. Splitting it into two deliveries would double-count the unit
    and misreport the course; collapsing the break away would lose four weeks of
    calendar.
    """

    unit_code: str
    start_date: dt.date
    end_date: dt.date
    #: Weeks the unit is actually taught.
    active_weeks: int
    #: Weeks from first to last teaching week inclusive, breaks included.
    calendar_weeks: int
    #: The break periods that fall inside this delivery.
    interruptions: list[Period] = field(default_factory=list)

    @property
    def is_interrupted(self) -> bool:
        return bool(self.interruptions)


@dataclass
class RollingIntake:
    """One intake column: when it starts and what it does, week by week."""

    #: Workbook column letter. Traceability only — never the business identity.
    source_column: str
    #: The week carrying the `Intake` marker. `None` for a stream already under
    #: way when the calendar opens.
    marker_week: CalendarWeek | None
    periods: list[Period]
    unit_deliveries: list[UnitDelivery]

    @property
    def intake_date(self) -> dt.date | None:
        """The intake's calendar date: the Monday of the marked week.

        Delivery begins the following week — the marker week is the entry point,
        not the first teaching week. Both facts come from the workbook: every
        marked column's first unit sits in the row directly below its marker.
        """
        return self.marker_week.start_date if self.marker_week else None

    @property
    def first_delivered_unit(self) -> str | None:
        return self.unit_deliveries[0].unit_code if self.unit_deliveries else None

    @property
    def start_date(self) -> dt.date | None:
        return self.periods[0].start_date if self.periods else None

    @property
    def end_date(self) -> dt.date | None:
        return self.periods[-1].end_date if self.periods else None

    @property
    def total_weeks(self) -> int:
        return sum(p.active_weeks for p in self.periods if p.period_type != INTAKE_MARKER)

    @property
    def teaching_weeks(self) -> int:
        return sum(p.active_weeks for p in self.periods if p.period_type == PERIOD_UNIT)

    @property
    def break_weeks(self) -> int:
        return sum(p.active_weeks for p in self.periods if p.period_type == PERIOD_BREAK)

    @property
    def assessment_weeks(self) -> int:
        return sum(p.active_weeks for p in self.periods if p.period_type == PERIOD_ASSESSMENT)


@dataclass
class RollingSheet:
    """A parsed qualification sheet."""

    qualification_code: str
    weeks: list[CalendarWeek]
    intakes: list[RollingIntake]
    source_workbook: str = ""

    @property
    def unit_codes(self) -> set[str]:
        return {d.unit_code for i in self.intakes for d in i.unit_deliveries}

    @property
    def calendar_start(self) -> dt.date:
        return self.weeks[0].start_date

    @property
    def calendar_end(self) -> dt.date:
        return self.weeks[-1].end_date


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def _as_date(value: object) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def _clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def _column_letter(index: int) -> str:
    """0-based column index to its spreadsheet letter."""
    letters = ""
    index += 1
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _locate_columns(header: Sequence[object]) -> tuple[int, int, int]:
    """Find the week/start/end columns by their headings, not by position."""
    found: dict[str, int] = {}
    for index, cell in enumerate(header):
        text = _clean(cell).lower()
        if not text:
            continue
        if _HEADER_WEEK in text and _HEADER_WEEK not in found:
            found[_HEADER_WEEK] = index
        elif text.startswith(_HEADER_START) and _HEADER_START not in found:
            found[_HEADER_START] = index
        elif text.startswith(_HEADER_END) and _HEADER_END not in found:
            found[_HEADER_END] = index

    missing = [k for k in (_HEADER_WEEK, _HEADER_START, _HEADER_END) if k not in found]
    if missing:
        raise RollingTimetableError(
            "The rolling timetable sheet must have Week No., Start Date and End Date "
            f"columns. Missing: {', '.join(missing)}."
        )
    return found[_HEADER_WEEK], found[_HEADER_START], found[_HEADER_END]


def _classify(value: str) -> tuple[str, str | None]:
    upper = value.upper()
    if upper == BREAK:
        return PERIOD_BREAK, None
    if upper == ASSESSMENT:
        return PERIOD_ASSESSMENT, None
    if upper == INTAKE_MARKER:
        return INTAKE_MARKER, None
    return PERIOD_UNIT, value.upper()


def _build_deliveries(periods: Sequence[Period]) -> list[UnitDelivery]:
    """Fold a column's periods into one delivery per continuous unit run.

    Two runs of the same unit separated only by breaks and assessment weeks are
    one interrupted delivery. Two runs separated by a *different unit* are two
    deliveries — the cycle has come round again — which is why this walks the
    periods in order rather than grouping by unit code.
    """
    deliveries: list[UnitDelivery] = []
    pending: UnitDelivery | None = None
    gap: list[Period] = []

    for period in periods:
        if period.period_type == INTAKE_MARKER:
            continue

        if period.period_type == PERIOD_UNIT:
            if pending is not None and pending.unit_code == period.value:
                # Same unit resuming after a break: extend rather than restart.
                pending.end_date = period.end_date
                pending.active_weeks += period.active_weeks
                pending.interruptions.extend(gap)
            else:
                if pending is not None:
                    deliveries.append(pending)
                pending = UnitDelivery(
                    unit_code=str(period.value),
                    start_date=period.start_date,
                    end_date=period.end_date,
                    active_weeks=period.active_weeks,
                    calendar_weeks=period.active_weeks,
                )
            gap = []
        else:
            # A break or assessment week only interrupts a delivery if the same
            # unit resumes after it; held aside until the next unit is known.
            gap.append(period)

    if pending is not None:
        deliveries.append(pending)

    for delivery in deliveries:
        span = (delivery.end_date - delivery.start_date).days + 1
        delivery.calendar_weeks = span // 7
    return deliveries


def parse_sheet(
    qualification_code: str,
    rows: Iterable[Sequence[object]],
    *,
    source_workbook: str = "",
) -> RollingSheet:
    """Parse one qualification's rolling sheet from its raw cell values.

    Takes rows rather than a file path so the parser can be tested without a
    workbook, and so the caller decides which sheet is in scope.
    """
    rows = [list(r) for r in rows]
    if not rows:
        raise RollingTimetableError("The rolling timetable sheet is empty.")

    week_col, start_col, end_col = _locate_columns(rows[0])
    fixed = {week_col, start_col, end_col}

    weeks: list[CalendarWeek] = []
    cells: list[dict[int, str]] = []

    for line_no, row in enumerate(rows[1:], start=2):
        if all(_clean(c) == "" for c in row):
            continue
        raw_week = row[week_col] if week_col < len(row) else None
        start = _as_date(row[start_col]) if start_col < len(row) else None
        end = _as_date(row[end_col]) if end_col < len(row) else None
        if raw_week is None or start is None or end is None:
            continue
        try:
            week_no = int(raw_week)
        except (TypeError, ValueError):
            raise RollingTimetableError(
                f"Row {line_no}: Week No. {raw_week!r} is not a whole number."
            ) from None
        if end < start:
            raise RollingTimetableError(
                f"Row {line_no}: End Date {end} is before Start Date {start}."
            )

        weeks.append(CalendarWeek(week_no=week_no, start_date=start, end_date=end))
        cells.append(
            {
                index: _clean(value)
                for index, value in enumerate(row)
                if index not in fixed and _clean(value)
            }
        )

    if not weeks:
        raise RollingTimetableError("The rolling timetable sheet has no dated week rows.")

    intake_columns = sorted({index for row in cells for index in row})
    intakes = [
        _build_intake(column, weeks, cells)
        for column in intake_columns
    ]

    return RollingSheet(
        qualification_code=qualification_code.upper(),
        weeks=weeks,
        intakes=intakes,
        source_workbook=source_workbook,
    )


def _build_intake(
    column: int, weeks: Sequence[CalendarWeek], cells: Sequence[dict[int, str]]
) -> RollingIntake:
    entries = [(week, row[column]) for week, row in zip(weeks, cells) if column in row]

    periods: list[Period] = []
    marker_week: CalendarWeek | None = None
    letter = _column_letter(column)

    for value, group in groupby(entries, key=lambda e: e[1].upper()):
        run = list(group)
        period_type, unit = _classify(run[0][1])
        if period_type == INTAKE_MARKER:
            marker_week = run[0][0]
            continue
        periods.append(
            Period(
                period_type=period_type,
                value=unit,
                first_week_no=run[0][0].week_no,
                last_week_no=run[-1][0].week_no,
                start_date=run[0][0].start_date,
                end_date=run[-1][0].end_date,
                active_weeks=len(run),
                source_column=letter,
            )
        )

    return RollingIntake(
        source_column=letter,
        marker_week=marker_week,
        periods=periods,
        unit_deliveries=_build_deliveries(periods),
    )


def load_workbook_sheet(path: str, sheet_name: str) -> RollingSheet:
    """Read one sheet from a rolling timetable workbook.

    Deliberately loads a single named sheet: the operational workbook holds a
    sheet per qualification, and reading all of them would pull data outside the
    approved scope of whatever import is running.
    """
    import openpyxl  # imported lazily: only an import run needs it

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RollingTimetableError(
                f"The workbook has no sheet named {sheet_name!r}. "
                f"Available: {', '.join(workbook.sheetnames)}."
            )
        sheet = workbook[sheet_name]
        return parse_sheet(
            sheet_name,
            sheet.iter_rows(values_only=True),
            source_workbook=path.rsplit("\\", 1)[-1],
        )
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Validation against approved TDMS data
# ---------------------------------------------------------------------------


@dataclass
class SheetValidation:
    """What a parsed sheet contains, checked against the approved units."""

    unknown_unit_codes: list[str]
    units_missing_from_sheet: list[str]
    intakes_without_marker: list[str]
    intakes_without_units: list[str]

    @property
    def is_consistent(self) -> bool:
        return not (self.unknown_unit_codes or self.units_missing_from_sheet)


def validate_against_units(sheet: RollingSheet, approved_unit_codes: Iterable[str]) -> SheetValidation:
    """Compare the sheet's unit codes with the qualification's approved units.

    The approved set comes from the database, never from the workbook, so a
    typo in a spreadsheet cannot quietly introduce a unit.
    """
    approved = {code.upper() for code in approved_unit_codes}
    found = sheet.unit_codes
    return SheetValidation(
        unknown_unit_codes=sorted(found - approved),
        units_missing_from_sheet=sorted(approved - found),
        intakes_without_marker=[i.source_column for i in sheet.intakes if i.marker_week is None],
        intakes_without_units=[i.source_column for i in sheet.intakes if not i.unit_deliveries],
    )
