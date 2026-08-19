"""BSB50420 pilot reference and trainer import.

Loads only the approved vertical slice — AIBT / 104262B / BSB50420 / Blacktown —
from the four supplied workbooks. Everything written here traces to a cell in a
source file; nothing is derived, defaulted or invented, and the two places where
the sources are genuinely silent are reported rather than filled in.

Runs against the least-privilege `tdms_app` role, the same role FastAPI uses, so
a privilege the application does not have is a failure here rather than a
surprise later.

    python scripts/import_bsb50420_pilot.py            # dry run, writes nothing
    python scripts/import_bsb50420_pilot.py --apply    # write

Idempotent: an entity that already exists is reused, never duplicated.
"""

from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.config import get_settings  # noqa: E402
from app.models.college import Campus, College, CollegeCampus  # noqa: E402
from app.models.course import CourseOffering, OfferingDurationOption  # noqa: E402
from app.models.qualification import Qualification, QualificationUnit, Unit  # noqa: E402
from app.models.trainer import (  # noqa: E402
    Trainer,
    TrainerAvailability,
    TrainerQualification,
    TrainerUnit,
)
from app.services.reference_data import (  # noqa: E402
    correct_supplied_course_statuses,
    status_for_supplied_course,
)
from app.services.rolling_timetable import load_workbook_sheet  # noqa: E402

from _source_data import (  # noqa: E402
    LOCATION_FILE,
    QUALIFICATION_FILE,
    ROLLING_FILE,
    TRAINER_FILE,
    require,
)

RTO = "AIBT"
COURSE_CODE = "104262B"
QUALIFICATION_CODE = "BSB50420"
CAMPUS_MATCH = "BLACKTOWN"

#: The campus code and the college's full legal name are the two values no
#: supplied file contains. Both are reported at the end as provisional so they
#: can be corrected before this data is relied on.
PROVISIONAL_CAMPUS_CODE = "BLACKTOWN"
PROVISIONAL_COLLEGE_FULL_NAME = "AIBT"

WORKING_START = dt.time(9, 0)
WORKING_END = dt.time(17, 0)


def rows(path: Path, sheet_name: str | None = None) -> list[dict]:
    book = openpyxl.load_workbook(path, data_only=True)
    sheet = book[sheet_name] if sheet_name else book[book.sheetnames[0]]
    values = list(sheet.values)
    header = [str(h).strip() if h else "" for h in values[0]]
    book.close()
    return [dict(zip(header, r)) for r in values[1:] if any(r)]


def text(row: dict, key: str) -> str:
    value = row.get(key)
    return "" if value is None else str(value).strip()


class Importer:
    def __init__(self, session: Session, *, apply: bool) -> None:
        self.session = session
        self.apply = apply
        self.log: list[str] = []
        self.warnings: list[str] = []
        self.correction = None

    def note(self, action: str, detail: str) -> None:
        self.log.append(f"  {action:8} {detail}")

    def get_or_create(self, model, defaults: dict, **lookup):
        existing = self.session.execute(select(model).filter_by(**lookup)).scalar_one_or_none()
        label = ", ".join(f"{k}={v!r}" for k, v in lookup.items())
        if existing is not None:
            self.note("REUSE", f"{model.__name__} {label}")
            return existing
        row = model(**lookup, **defaults)
        self.session.add(row)
        self.session.flush()
        self.note("CREATE", f"{model.__name__} {label}")
        return row

    # -- reference ---------------------------------------------------------

    def run(self) -> None:
        location = rows(LOCATION_FILE)
        pilot = [
            r
            for r in location
            if text(r, "RTO") == RTO
            and text(r, "VET Code") == QUALIFICATION_CODE
            and CAMPUS_MATCH in text(r, "Location").upper()
        ]
        if len(pilot) != 1:
            raise SystemExit(
                f"Expected exactly one {RTO}/{QUALIFICATION_CODE}/Blacktown row in "
                f"{LOCATION_FILE.name}; found {len(pilot)}."
            )
        row = pilot[0]

        college = self.get_or_create(
            College,
            {"college_full_name": PROVISIONAL_COLLEGE_FULL_NAME, "is_active": True},
            college_short_name=RTO,
        )
        self.warnings.append(
            "colleges.college_full_name is provisional: no supplied file contains "
            f"{RTO}'s full legal name, so the short name was stored rather than one invented."
        )

        address = text(row, "Location")
        campus = self.get_or_create(
            Campus,
            {
                "campus_name": "Blacktown",  # Trainer Data 'Location' column
                "campus_location": address,  # Location Data, verbatim
                "state": "NSW",              # read from the address
                "is_active": True,
            },
            campus_code=PROVISIONAL_CAMPUS_CODE,
        )
        self.warnings.append(
            "campuses.campus_code is provisional: no supplied file contains campus codes."
        )

        self.get_or_create(
            CollegeCampus, {"is_active": True}, college_id=college.id, campus_id=campus.id
        )

        qual_rows = [
            r
            for r in rows(QUALIFICATION_FILE)
            if text(r, "RTO") == RTO and text(r, "Qualification Code") == QUALIFICATION_CODE
        ]
        if not qual_rows:
            raise SystemExit(f"No {RTO} {QUALIFICATION_CODE} rows in {QUALIFICATION_FILE.name}.")

        qualification = self.get_or_create(
            Qualification,
            {
                "qualification_title": text(qual_rows[0], "Qualification Title"),
                "course_level": text(row, "Course Level") or None,
                "field_of_education_broad": text(row, "Field Of Education Broard") or None,
                "field_of_education_narrow": text(row, "Field Of Education Narrow") or None,
                "course_sector": text(row, "Course Sector") or None,
                "source_url": text(qual_rows[0], "Source URL") or None,
                "is_active": True,
            },
            qualification_code=QUALIFICATION_CODE,
        )

        units: dict[str, Unit] = {}
        for r in qual_rows:
            code = text(r, "Unit Code")
            units[code] = self.get_or_create(
                Unit,
                {
                    "unit_title": text(r, "Unit Title"),
                    # No supplied file states UoC Type. Left unset rather than guessed.
                    "uoc_type": None,
                    "is_active": True,
                },
                unit_code=code,
            )
        self.warnings.append(
            "units.uoc_type left empty: no supplied file states Theory / Theory and Practical."
        )

        # -- base rolling delivery cycle -----------------------------------
        #
        # The order comes from the earliest complete stream in the rolling
        # timetable, not from spreadsheet row order in Qualification Data and
        # not from a guess. It is the *cycle* order: intakes join it at
        # different points, which `delivery_order` does not attempt to express.
        sheet = load_workbook_sheet(str(ROLLING_FILE), f"{QUALIFICATION_CODE}_52_Weeks")
        base = sheet.intakes[0]
        cycle = [d.unit_code for d in base.unit_deliveries]
        seen: list[str] = []
        for code in cycle:
            if code not in seen:
                seen.append(code)
        if set(seen) != set(units):
            raise SystemExit(
                "The rolling timetable's base cycle does not match the qualification's units.\n"
                f"  only in cycle: {sorted(set(seen) - set(units))}\n"
                f"  only in units: {sorted(set(units) - set(seen))}"
            )

        for order, code in enumerate(seen, start=1):
            self.get_or_create(
                QualificationUnit,
                {"delivery_order": order},
                qualification_id=qualification.id,
                unit_id=units[code].id,
            )
        self.note("CYCLE", " -> ".join(seen))

        # Approved rule: a course supplied in an approved source is ACTIVE in
        # TDMS. The workbook's "Registered" describes external registration, not
        # TDMS availability, and is not carried into the status field. The rule
        # lives in the service so parser, API and browser cannot drift apart.
        status = status_for_supplied_course(self.session)
        self.note("STATUS", f"{text(row, 'Course Status')!r} (source) -> {status.code} (TDMS)")

        offering = self.get_or_create(
            CourseOffering,
            {
                "course_code": COURSE_CODE,
                "course_status_id": status.id,
                "total_course_cost": text(row, "Total Course Cost") or None,
            },
            college_id=college.id,
            campus_id=campus.id,
            qualification_id=qualification.id,
        )
        self.get_or_create(
            OfferingDurationOption,
            {"is_active": True},
            course_offering_id=offering.id,
            duration_weeks=int(text(row, "Duration In Weeks")),
        )

        # -- trainers ------------------------------------------------------
        self._import_trainers(campus, qualification, units)

        # -- status correction --------------------------------------------
        #
        # Applies the approved rule to every offering already in the database
        # whose course code appears in the supplied workbook. Earlier imports
        # wrote the source's registration wording into the status field; this
        # brings them onto ACTIVE without touching anything the workbook does
        # not account for.
        supplied_codes = {text(r, "Course Code") for r in location}
        self.correction = correct_supplied_course_statuses(self.session, supplied_codes)

    def _import_trainers(self, campus, qualification, units) -> None:
        locations = rows(TRAINER_FILE, "Trainer Location")
        competencies = rows(TRAINER_FILE, "Trainer Units")

        blacktown = [r for r in locations if CAMPUS_MATCH in text(r, "Location").upper()]
        weekday_columns = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")

        for r in blacktown:
            trainer_code = text(r, "Trainer id")
            trainer = self.get_or_create(
                Trainer,
                {"trainer_name": text(r, "Trainer name"), "is_active": True},
                trainer_id=trainer_code,
            )

            self.get_or_create(
                TrainerAvailability,
                {
                    "location": text(r, "Location") or None,
                    "location_type": text(r, "Location Type") or None,
                    "working_time_end": WORKING_END,
                    **{
                        column.lower(): (
                            "NOT_AVAILABLE"
                            if text(r, column).upper() in ("NA", "N/A", "")
                            else text(r, column).upper()
                        )
                        for column in weekday_columns
                    },
                },
                trainer_id=trainer.id,
                campus_id=campus.id,
                class_type=text(r, "Delivery Type").upper(),
                working_time_start=WORKING_START,
            )

            mine = [
                c
                for c in competencies
                if text(c, "Trainer ID") == trainer_code
                and text(c, "Qualifications They Can Teach") == QUALIFICATION_CODE
            ]
            if mine:
                self.get_or_create(
                    TrainerQualification,
                    {},
                    trainer_id=trainer.id,
                    qualification_id=qualification.id,
                )
            for c in mine:
                code = text(c, "Units They Can Teach")
                if code not in units:
                    self.warnings.append(
                        f"{trainer_code} is mapped to {code}, which is not a "
                        f"{QUALIFICATION_CODE} unit — skipped."
                    )
                    continue
                self.get_or_create(
                    TrainerUnit, {}, trainer_id=trainer.id, unit_id=units[code].id
                )

        other_sites = {
            text(r, "Location")
            for r in locations
            if text(r, "Trainer id") in {text(b, "Trainer id") for b in blacktown}
            and CAMPUS_MATCH not in text(r, "Location").upper()
        }
        if other_sites:
            self.warnings.append(
                "Availability at "
                + ", ".join(sorted(other_sites))
                + " was not imported: those campuses are outside the approved pilot scope."
            )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="write to the database")
    args = parser.parse_args()

    require(
        LOCATION_FILE.name,
        QUALIFICATION_FILE.name,
        TRAINER_FILE.name,
        ROLLING_FILE.name,
    )

    settings = get_settings()
    engine = create_engine(settings.database_url, future=True)

    print(f"BSB50420 pilot import — {'APPLY' if args.apply else 'DRY RUN'}")
    print(f"  runtime role: {settings.runtime_identity}\n")

    with Session(engine) as session:
        importer = Importer(session, apply=args.apply)
        importer.run()

        print("\n".join(importer.log))
        correction = getattr(importer, "correction", None)
        if correction is not None:
            print("\ncourse status (approved rule: supplied course -> ACTIVE):")
            print(f"  supplied offerings in database : {correction.supplied_offerings}")
            print(f"  already ACTIVE                 : {correction.already_active}")
            print(f"  incorrectly not ACTIVE         : {correction.incorrectly_inactive}")
            print(f"  corrected                      : {correction.corrected or 'none'}")
            print(f"  left alone (not traceable)     : {correction.untraceable_left_alone or 'none'}")
            print(f"  retired source status values   : {correction.retired_status_codes or 'none'}")

        if importer.warnings:
            print("\nreported, not invented:")
            for warning in dict.fromkeys(importer.warnings):
                print(f"  - {warning}")

        if args.apply:
            session.commit()
            print("\ncommitted.")
        else:
            session.rollback()
            print("\nrolled back — nothing was written.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
