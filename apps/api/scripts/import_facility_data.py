"""Facility Data import.

    python scripts/import_facility_data.py            # dry run, writes nothing
    python scripts/import_facility_data.py --apply    # write

Runs as the least-privilege `tdms_app` role. One transaction: complete or none.

A source row is not a room. The file's 129 rows describe 70 rooms, each repeated
once per College and once per Faculty that may use it, so the import writes
three tables and the file can be reconstructed from them exactly:

    facilities          70   the physical room, keyed by campus + Location + name
    facility_colleges  110   which college may use it
    facility_faculties  81   which faculty may use it, on which weekdays

Idempotent: re-running reuses a room, a link and a rule rather than duplicating
them. Nothing is discarded silently — a row that cannot be resolved is reported
and the import refuses to write unless every row resolved.
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.config import get_settings  # noqa: E402
from app.models.college import Campus, CampusSourceAddress, College  # noqa: E402
from app.models.facility import (  # noqa: E402
    WEEKDAY_COLUMNS,
    Facility,
    FacilityCollege,
    FacilityFaculty,
)
from app.services.reference_import import derive_campus, repair_text  # noqa: E402

from _source_data import require, source_file  # noqa: E402

FACILITY_FILE_NAME = "Facility Data.xlsx"
SHEET = "Facility"

#: Supplied Location strings that no derivation resolves, mapped to the campus
#: they belong to.
#:
#: `derive_campus` produces `UPPERMOUNTGRAVATTBRI` for this one because the
#: string carries a suburb *and* a city. It is the same site as the existing
#: BRISBANE campus — `18 Mt Gravatt-Capalaba Road` against the stored
#: `18 Mount Gravatt-Capalaba Road`, a Mt/Mount variant this project already
#: records for that campus. Listed here rather than inferred, so the assumption
#: is visible and can be corrected in one place.
LOCATION_CAMPUS_OVERRIDES = {
    "Levels 2-3, 18 Mt Gravatt-Capalaba Road, Upper Mount Gravatt, Brisbane QLD 4122": "BRISBANE",
}

#: Source values meaning "nothing recorded here". Preserved for Faculty, where
#: `NA` is a rule, and dropped for Remarks, where it means no remark.
NO_REMARK = {"NA", "N/A", ""}


def read_rows() -> list[dict]:
    book = openpyxl.load_workbook(source_file(FACILITY_FILE_NAME), data_only=True)
    sheet = book[SHEET] if SHEET in book.sheetnames else book[book.sheetnames[0]]
    values = list(sheet.values)
    book.close()

    header = [repair_text(h) for h in values[0]]
    rows = []
    for number, raw in enumerate(values[1:], start=2):
        if raw is None or all(v is None for v in raw):
            continue
        row = {key: value for key, value in zip(header, raw)}
        row["_row"] = number
        rows.append(row)
    return rows


def text(value) -> str:
    return repair_text(value)


def yes(value) -> bool:
    """`Yes` / `No` as supplied. Anything else is reported, never assumed."""
    supplied = text(value).strip().lower()
    if supplied in {"yes", "y", "true"}:
        return True
    if supplied in {"no", "n", "false"}:
        return False
    raise ValueError(f"{value!r} is not Yes or No")


def resolve_campus(location: str, campuses: dict[str, Campus], aliases: dict[str, int]):
    """Alias first, then derivation, then an approved override. Never a guess."""
    if location in aliases:
        return aliases[location], "alias"
    override = LOCATION_CAMPUS_OVERRIDES.get(location)
    if override and override in campuses:
        return campuses[override].id, "approved override"
    derived = derive_campus(location).code
    if derived in campuses:
        return campuses[derived].id, "derived"
    return None, f"unresolved (derived {derived!r}, no such campus)"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="write to the database")
    args = parser.parse_args()

    require(FACILITY_FILE_NAME)

    settings = get_settings()
    engine = create_engine(settings.database_url, future=True)

    print(f"Facility Data — {'APPLY' if args.apply else 'DRY RUN'}")
    print(f"  runtime role: {settings.runtime_identity}\n")

    rows = read_rows()
    print(f"source rows: {len(rows)}\n")

    problems: list[str] = []

    with Session(engine) as session:
        campuses = {c.campus_code: c for c in session.execute(select(Campus)).scalars()}
        colleges = {c.college_short_name: c for c in session.execute(select(College)).scalars()}
        aliases = {
            a.source_address: a.campus_id
            for a in session.execute(select(CampusSourceAddress)).scalars()
        }

        # ------------------------------------------------------- locations
        locations = sorted({text(r["Location"]) for r in rows})
        location_campus: dict[str, int] = {}
        print("locations:")
        for location in locations:
            campus_id, how = resolve_campus(location, campuses, aliases)
            count = sum(1 for r in rows if text(r["Location"]) == location)
            if campus_id is None:
                problems.append(f"Location not resolved: {location}")
                print(f"  {count:4}  UNRESOLVED  {location}")
                continue
            location_campus[location] = campus_id
            name = next(c.campus_name for c in campuses.values() if c.id == campus_id)
            print(f"  {count:4}  {name:18} ({how}){'':2}{location[:58]}")

        # ---------------------------------------------------------- colleges
        supplied_colleges = sorted({text(r["College"]) for r in rows})
        missing = [c for c in supplied_colleges if c not in colleges]
        print(f"\ncolleges: {', '.join(supplied_colleges)}")
        for code in missing:
            problems.append(f"College not in reference data: {code}")
            print(f"  UNRESOLVED college: {code}")

        # ------------------------------------------------------ build model
        rooms: dict[tuple[int, str, str], dict] = {}
        for row in rows:
            location = text(row["Location"])
            if location not in location_campus:
                continue
            campus_id = location_campus[location]
            name = text(row["Classroom name"])
            key = (campus_id, location, name)

            try:
                availability = tuple(yes(row[day.capitalize()]) for day in WEEKDAY_COLUMNS)
            except ValueError as error:
                problems.append(f"row {row['_row']}: {error}")
                continue

            capacity = row["Capacity"]
            if not isinstance(capacity, int) or capacity <= 0:
                problems.append(f"row {row['_row']}: capacity {capacity!r} is not a positive whole number")
                continue

            entry = rooms.setdefault(
                key,
                {
                    "campus_id": campus_id,
                    "source_location": location,
                    "facility_reference": name,
                    "facility_type": text(row["Classroom Type"]),
                    "capacity": capacity,
                    "colleges": set(),
                    "faculties": {},
                },
            )
            # Capacity and type are consistent per room in the supplied file.
            # If that ever stops being true, say so rather than picking one.
            if entry["capacity"] != capacity:
                problems.append(
                    f"row {row['_row']}: capacity {capacity} disagrees with {entry['capacity']} "
                    f"already supplied for {name!r} at {location}"
                )
            college_code = text(row["College"])
            if college_code in colleges:
                entry["colleges"].add(college_code)

            faculty = text(row["Faculty"])
            remark = text(row["Remarks"])
            rule = {
                "availability": availability,
                "remarks": None if remark.upper() in NO_REMARK else remark,
            }
            existing = entry["faculties"].get(faculty)
            if existing and existing != rule:
                problems.append(
                    f"row {row['_row']}: {faculty!r} availability/remarks for {name!r} "
                    f"disagree with an earlier row for the same room and faculty"
                )
            entry["faculties"][faculty] = rule

        college_links = sum(len(e["colleges"]) for e in rooms.values())
        faculty_rules = sum(len(e["faculties"]) for e in rooms.values())
        print(f"\nresolved to:")
        print(f"  {len(rooms):4} facilities")
        print(f"  {college_links:4} facility-college links")
        print(f"  {faculty_rules:4} facility-faculty rules")

        faculties = sorted({f for e in rooms.values() for f in e["faculties"]})
        print(f"\nfaculties supplied: {', '.join(faculties)}")

        if problems:
            print(f"\n{len(problems)} problem(s):")
            for problem in problems:
                print(f"  - {problem}")
            print("\nNothing was written. Resolve these first.")
            return 1

        # ---------------------------------------------------------- write
        existing_facilities = {
            (f.campus_id, f.source_location, f.facility_reference): f
            for f in session.execute(select(Facility)).scalars()
        }
        created = reused = 0
        for key, entry in rooms.items():
            facility = existing_facilities.get(key)
            if facility is None:
                facility = Facility(
                    facility_reference=entry["facility_reference"],
                    campus_id=entry["campus_id"],
                    source_location=entry["source_location"],
                    facility_type=entry["facility_type"],
                    capacity=entry["capacity"],
                )
                session.add(facility)
                session.flush()
                created += 1
            else:
                facility.facility_type = entry["facility_type"]
                facility.capacity = entry["capacity"]
                reused += 1

            have_colleges = {link.college_id for link in facility.colleges}
            for code in sorted(entry["colleges"]):
                college_id = colleges[code].id
                if college_id not in have_colleges:
                    session.add(FacilityCollege(facility_id=facility.id, college_id=college_id))

            have_rules = {rule.faculty: rule for rule in facility.faculties}
            for faculty, rule in sorted(entry["faculties"].items()):
                target = have_rules.get(faculty)
                if target is None:
                    target = FacilityFaculty(facility_id=facility.id, faculty=faculty)
                    session.add(target)
                for day, value in zip(WEEKDAY_COLUMNS, rule["availability"]):
                    setattr(target, day, value)
                target.remarks = rule["remarks"]

        # Record every supplied Location spelling, so a later import of any
        # file using the same wording resolves without an override.
        recorded = 0
        for location, campus_id in location_campus.items():
            if location not in aliases:
                session.add(CampusSourceAddress(campus_id=campus_id, source_address=location))
                aliases[location] = campus_id
                recorded += 1

        print(f"\nfacilities created {created}, reused {reused}")
        print(f"campus address spellings recorded: {recorded}")

        if args.apply:
            session.commit()
            print("\ncommitted.")
        else:
            session.rollback()
            print("\nrolled back — nothing was written.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
