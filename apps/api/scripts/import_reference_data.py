"""Full College and Course Reference Data import.

    python scripts/import_reference_data.py            # dry run, writes nothing
    python scripts/import_reference_data.py --apply    # write

Runs as the least-privilege `tdms_app` role — the same role FastAPI uses — so a
privilege the application lacks fails here rather than in production. The whole
import is one transaction: it lands complete or not at all.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.config import get_settings  # noqa: E402
from app.services.reference_import import ReferenceImporter, format_report  # noqa: E402
from app.services.rolling_timetable import load_workbook_sheet  # noqa: E402

from _source_data import (  # noqa: E402
    LOCATION_FILE,
    QUALIFICATION_FILE,
    ROLLING_FILE,
    require,
)

#: Qualifications with an approved rolling timetable. Only these get a stored
#: delivery order; everything else keeps membership pending rather than being
#: given an invented 1..N sequence.
ROLLING_SHEETS = {"BSB50420": "BSB50420_52_Weeks"}


def read(path: Path) -> list[dict]:
    book = openpyxl.load_workbook(path, data_only=True)
    sheet = book[book.sheetnames[0]]
    values = list(sheet.values)
    header = [str(h).strip() if h else "" for h in values[0]]
    book.close()
    return [
        dict(zip(header, row))
        for row in values[1:]
        if any(cell is not None and str(cell).strip() for cell in row)
    ]


def base_cycles() -> dict[str, list[str]]:
    """Approved delivery order per qualification, read from the rolling timetable.

    The order is the earliest complete stream's unit progression — the base
    rolling cycle. It is evidence, not an assumption, and intakes join it at
    different points, which `delivery_order` does not attempt to express.
    """
    cycles: dict[str, list[str]] = {}
    for code, sheet_name in ROLLING_SHEETS.items():
        sheet = load_workbook_sheet(str(ROLLING_FILE), sheet_name)
        seen: list[str] = []
        for delivery in sheet.intakes[0].unit_deliveries:
            if delivery.unit_code not in seen:
                seen.append(delivery.unit_code)
        cycles[code] = seen
    return cycles


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="write to the database")
    args = parser.parse_args()

    require(LOCATION_FILE.name, QUALIFICATION_FILE.name, ROLLING_FILE.name)

    settings = get_settings()
    engine = create_engine(settings.database_url, future=True)

    location_rows = read(LOCATION_FILE)
    qualification_rows = read(QUALIFICATION_FILE)
    cycles = base_cycles()

    print(f"College and Course Reference Data import — {'APPLY' if args.apply else 'DRY RUN'}")
    print(f"  runtime role      : {settings.runtime_identity}")
    print(f"  {LOCATION_FILE.name:32} {len(location_rows)} rows")
    print(f"  {QUALIFICATION_FILE.name:32} {len(qualification_rows)} rows")
    for code, order in cycles.items():
        print(f"  approved sequence source        : {code} ({len(order)} units, from rolling timetable)")

    with Session(engine) as session:
        importer = ReferenceImporter(session, sequence_sources=cycles)
        report = importer.run(location_rows, qualification_rows)
        print(format_report(report))

        if args.apply:
            session.commit()
            print("\ncommitted.")
        else:
            session.rollback()
            print("\nrolled back — nothing was written.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
